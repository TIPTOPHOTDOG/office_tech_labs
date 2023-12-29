import sys
from PyQt5.QtWidgets import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
from datetime import datetime, timedelta
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def warning() -> None:
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setText("Fill in all the fields")
    msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    msg_box.exec()


class MyWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.workbook = None
        self.students = None
        self.file_path = None
        self.processed_lab_sheets = set()
        self.init_ui()

    def init_ui(self) -> None:
        create = QPushButton("Create journal")
        add_lab = QPushButton("Add lab")
        cental_widget = QWidget()
        self.create_label = QLabel()
        self.create_label.setFixedSize(300, 20)
        self.lab_count = QLabel("0")
        self.lab_count.setFixedSize(300, 20)
        self.lab_count.setStyleSheet("background: white")
        self.setCentralWidget(cental_widget)
        main_layout = QVBoxLayout()
        hbox = QHBoxLayout()
        hbox.addWidget(create)
        hbox.addWidget(self.create_label)
        main_layout.addLayout(hbox)
        hbox = QHBoxLayout()
        hbox.addWidget(add_lab)
        hbox.addWidget(self.lab_count)
        create.clicked.connect(self.create_journal)
        add_lab.clicked.connect(self.add_lab)
        main_layout.addLayout(hbox)
        cental_widget.setLayout(main_layout)
        self.setFixedSize(420, 75)

    def create_journal(self) -> None:
        journal_dialog = QDialog()
        main_layout = QVBoxLayout()
        hbox = QHBoxLayout()
        open_students_file_button = QPushButton("Open students file")
        self.students_label = QLabel("")
        hbox.addWidget(open_students_file_button)
        open_students_file_button.clicked.connect(self.open_students_file)
        self.start_date = QDateEdit()
        self.num_lectures = QLineEdit("num lectors")
        main_layout.addWidget(self.num_lectures)
        hbox.addWidget(self.students_label)
        main_layout.addWidget(self.start_date)
        main_layout.addLayout(hbox)
        hbox = QHBoxLayout()
        create_excel = QPushButton("Create excel journal")
        create_excel.clicked.connect(self.save_to_excel)
        self.excel_name = QLineEdit("output filename")
        main_layout.addWidget(self.excel_name)
        main_layout.addWidget(create_excel)
        journal_dialog.setLayout(main_layout)
        journal_dialog.exec()

    def open_students_file(self) -> None:
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(self,
                                                   'Open File ',
                                                   '',
                                                   'Text File (*.txt)',
                                                   options=options)
        if file_name:
            with open(file_name) as f:
                self.students = [i for i in f]
            self.students_label.setText(file_name)

    def save_to_excel(self) -> None:
        if not self.num_lectures.text() or not self.students_label.text() or not self.excel_name.text():
            Warning()
            return
        else:
            try:
                i = int(self.num_lectures.text())
            except Exception as e:
                Warning()
                return

        num_lectures = int(self.num_lectures.text())
        self.file_path = self.excel_name.text() + '.xlsx'
        start_date_str = self.start_date.date().toString()
        date_array = [i for i in start_date_str.split()]
        mounth = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                  "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
                  "янв": 1, "февр": 2, "март": 3, "апр": 4, "май": 5, "июнь": 6,
                  "июль": 7, "авг": 8, "сент": 9, "октб": 10, "нояб": 11, "дек": 12}
        start_date = datetime.strptime(f"{date_array[3]}-{mounth[date_array[1]]}-{date_array[2]}", "%Y-%m-%d")
        dates = [start_date + timedelta(days=7 * i) for i in range(num_lectures)]

        self.workbook = Workbook()

        attendance_sheet = self.workbook.create_sheet(title="Посещаемость", index=0)
        del self.workbook['Sheet']

        # Clear data in the new sheets
        attendance_sheet.delete_cols(2, attendance_sheet.max_column)

        for sheet in [attendance_sheet]:
            sheet.append(["Студент"] + [date.strftime('%Y-%m-%d') for date in dates] + ["Процент явки"])
            for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=num_lectures + 2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
            sheet.column_dimensions['A'].width = 45

            scores_sheet = self.workbook.create_sheet(title="Баллы", index=1)
            scores_sheet.delete_cols(2, scores_sheet.max_column)

            for sheet in [scores_sheet]:
                sheet.append(["Студент"])
                scores_sheet['E1'] = "Посещаемость 0/5"
                scores_sheet['C1'] = "Зачёт"
                scores_sheet['C1'].font = Font(bold=True)
                scores_sheet['D1'] = "За семестр"
                scores_sheet['D1'].font = Font(bold=True)
                scores_sheet['B1'] = "Итого"
                scores_sheet['B1'].font = Font(bold=True)

                for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center')
                    sheet.column_dimensions['A'].width = 45
                    sheet.column_dimensions['B'].width = 15

        for student in self.students:
            attendance_sheet.append([student] + [""] * num_lectures)
            scores_sheet.append([student])

        column_letter = str(attendance_sheet.cell(row=1, column=num_lectures + 1).column_letter)
        for row_num in range(2, len(self.students) + 2):
            scores_sheet[f'D{row_num}'] = f'=SUM(E{row_num}:X{row_num})'
            scores_sheet[f'B{row_num}'] = f'=SUM(C{row_num}:D{row_num})'

        for row_num in range(2, len(self.students) + 2):
            formula = f'=COUNTIF(B{row_num}:{column_letter}{row_num}, "<>-") / {num_lectures} * 100'
            attendance_sheet.cell(row=row_num, column=num_lectures + 2).value = formula

        red_background = PatternFill(fill_type='solid', bgColor="00FF0000")
        fill_style_red = DifferentialStyle(fill=red_background)

        yellow_background = PatternFill(fill_type='solid', bgColor="FFFF00")
        fill_style_yellow = DifferentialStyle(fill=yellow_background)

        rule_red = Rule(type="beginsWith", dxf=fill_style_red)
        rule_red.formula = ['NOT(ISERROR(SEARCH("-", B2)))']

        rule_yellow = Rule(type="containsText", dxf=fill_style_yellow)
        rule_yellow.formula = ['AND(NOT(ISERROR(SEARCH("-", B2))),LEN(B2)>1)']

        color_scale_rule = ColorScaleRule(start_type="num", start_value=0, start_color="00FF0000", mid_type="num",
                                          mid_value=50, mid_color="00FFFF00", end_type="num", end_value=100,
                                          end_color="0000FF00")

        attendance_sheet.conditional_formatting.add(f"B2:{column_letter}{len(self.students) + 1}", rule_yellow)
        attendance_sheet.conditional_formatting.add(f"B2:{column_letter}{len(self.students) + 1}", rule_red)
        column_letter = str(attendance_sheet.cell(row=1, column=num_lectures + 2).column_letter)
        attendance_sheet.conditional_formatting.add(f"{column_letter}2:{column_letter}{len(self.students) + 1}",
                                                    color_scale_rule)

        color_scale_rule = ColorScaleRule(start_type="num", start_value=0, start_color="00FF0000", mid_type="num",
                                          mid_value=30, mid_color="00FFFF00", end_type="num", end_value=60,
                                          end_color="0000FF00")
        column_letter = str(scores_sheet.cell(row=1, column=4).column_letter)
        scores_sheet.conditional_formatting.add(f"{column_letter}2:{column_letter}{len(self.students) + 1}",
                                                color_scale_rule)

        color_scale_rule = ColorScaleRule(start_type="num", start_value=0, start_color="00FF0000", mid_type="num",
                                          mid_value=20, mid_color="00FFFF00", end_type="num", end_value=40,
                                          end_color="0000FF00")
        column_letter = str(scores_sheet.cell(row=1, column=3).column_letter)
        scores_sheet.conditional_formatting.add(f"{column_letter}2:{column_letter}{len(self.students) + 1}",
                                                color_scale_rule)

        color_scale_rule = ColorScaleRule(start_type="num", start_value=0, start_color="00FF0000", mid_type="num",
                                          mid_value=50, mid_color="00FFFF00", end_type="num", end_value=100,
                                          end_color="0000FF00")
        column_letter = str(scores_sheet.cell(row=1, column=2).column_letter)
        scores_sheet.conditional_formatting.add(f"{column_letter}2:{column_letter}{len(self.students) + 1}",
                                                color_scale_rule)

        last_column = attendance_sheet.max_column
        last_column_letter = get_column_letter(last_column)

        formula_str = f'=Посещаемость!{last_column_letter}2:Посещаемость!{last_column_letter}{attendance_sheet.max_row}'

        for row_num in range(2, len(self.students) + 2):
            scores_sheet[f'E{row_num}'] = formula_str
            scores_sheet[f'E{row_num}'] = f'=Посещаемость!{last_column_letter}{row_num}*5/100'

        for i in range(2, 2 + num_lectures):
            letter = get_column_letter(i)
            row_range = f'{letter}2:{letter}{len(self.students) + 1}'
            formula = f'=COUNTBLANK({row_range})'
            attendance_sheet[f'{letter}{len(self.students) + 2}'].value = formula

        self.create_label.setText(self.file_path)
        self.workbook.save(self.file_path)

    def add_lab(self) -> None:
        if not self.file_path:
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText("Before adding a lab, create a journal")
            msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msg_box.exec()
            return
        add_lab_dialog = QDialog()
        labels = ['title of laboratory work',
                  'name of the first criterion',
                  'maximum score for the first criterion',
                  'name of the second criterion',
                  'maximum score for the second criterion']

        labels = [QLabel(i) for i in labels]
        self.line_edits = [QLineEdit() for i in labels]

        main_layout = QVBoxLayout()
        for i in range(len(labels)):
            hbox = QHBoxLayout()
            hbox.addWidget(labels[i])
            labels[i].setFixedSize(250, 20)
            hbox.addWidget(self.line_edits[i])
            self.line_edits[i].setFixedSize(250, 20)
            main_layout.addLayout(hbox)
        save_button = QPushButton("Add lab")
        save_button.clicked.connect(self.save_lab)
        main_layout.addWidget(save_button)
        add_lab_dialog.setLayout(main_layout)
        add_lab_dialog.exec()

    def save_lab(self) -> None:
        if not all(self.line_edits):
            warning()
            return
        lab_name = self.line_edits[0].text()
        self.lab_count.setText(str(int(self.lab_count.text()) + 1))
        print(lab_name, type(lab_name))
        lab_sheet = self.workbook.create_sheet(title=lab_name, index=1)

        lab_sheet.append(
            ["Студент", f"{self.line_edits[1].text() } ({self.line_edits[2].text() } б.)",
             f"{self.line_edits[3].text() } ({self.line_edits[4].text() } б.)",
             "Итого"])

        for student in self.students:
            lab_sheet.append([student, "", ""])

        for row_num in range(2, len(self.students) + 2):
            total_cell = f'D{row_num}'
            total_formula = f'=B{row_num} + C{row_num}'
            lab_sheet[total_cell].value = total_formula

            color_scale_rule = ColorScaleRule(start_type="num", start_value=0, start_color="FF0000", end_type="num",
                                              end_value=float(self.line_edits[2].text() ) + float(self.line_edits[4].text() ),
                                              end_color="00FF00")
            lab_sheet.conditional_formatting.add(total_cell, color_scale_rule)
        lab_sheet.column_dimensions['A'].width = 45
        lab_sheet.column_dimensions['B'].width = 25
        lab_sheet.column_dimensions['C'].width = 25

        self.workbook.save(self.file_path)
        self.add_lab_to_scores_sheet(lab_name)

    def add_lab_to_scores_sheet(self, lab_name) -> None:
        scores_sheet = self.workbook['Баллы']
        if lab_name in self.processed_lab_sheets:
            return
        else:
            self.processed_lab_sheets.add(lab_name)

        if lab_name in scores_sheet[1]:
            return
        else:
            new_column = scores_sheet.max_column + 1
            new_column_letter = get_column_letter(new_column)

            scores_sheet[f'{new_column_letter}1'] = lab_name

        lab_column = f'{lab_name}!D'
        lab_start_cell = f'{lab_name}!D2'

        formula_str = f'={lab_name}!D2:{lab_name}!D{len(self.students)+1}'

        for row_num in range(2, len(scores_sheet['A']) + 1):
            scores_sheet[f'{new_column_letter}{row_num}'] = formula_str
        self.workbook.save(self.file_path)

        workbook = load_workbook(self.file_path)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
