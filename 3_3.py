import sys
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, AreaChart
from PyQt5.QtWidgets import *

def warning() -> None:
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Warning)
    msg_box.setText("Fill in all the fields")
    msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    msg_box.exec()

class MyWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.file_path = ''
        self.init_ui()

    def init_ui(self) -> None:
        central_widget = QWidget()
        main_layout = QVBoxLayout()
        open_xls = QPushButton("Open xlsx")
        draw = QPushButton("Draw")
        draw.clicked.connect(self.draw)
        main_layout.addWidget(open_xls)
        main_layout.addWidget(draw)
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def open_xls(self) -> None:
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(self,
                                                   'Open Excel File ',
                                                   '',
                                                   'Excel File (*.xlsx)',
                                                   options=options)
        if file_name:
            self.file_path = file_name
            self.search_label.setText(file_name)

    def read_data(self, sheet) -> list:
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data


    def create_chart(self, title, chart_type, data, labels, new_sheet, start_row) -> None:
        chart = chart_type()
        chart.title = title
        chart.x_axis.title = 'ФИО'
        chart.y_axis.title = 'Значение'

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        new_sheet.add_chart(chart, f"A{start_row}")

    def draw(self) -> None:
        if not self.file_path:
            warning()
        else:
            # Загрузка существующей книги
            workbook = load_workbook(self.file_path)

            # Создание нового листа для графиков
            graph_sheet = workbook.create_sheet(title="Графики")

            # Считываем данные с существующих листов
            data_sheet1 = self.read_data(workbook["Посещаемость"])
            data_sheet2 = self.read_data(workbook["лр1"])
            data_sheet3 = self.read_data(workbook["Баллы"])

            # Строим графики на новом листе
            start_row = 1
            chart_title = "График 1"
            chart_type = BarChart
            data_reference = Reference(workbook["Посещаемость"], min_col=17, min_row=1, max_row=len(data_sheet1) + 1)
            categories_reference = Reference(workbook["Посещаемость"], min_col=1, min_row=2, max_row=len(data_sheet1) + 1)

            chart = chart_type()
            chart.title = chart_title
            chart.x_axis.title = 'ФИО'
            chart.y_axis.title = 'Значение'
            chart.add_data(data_reference, titles_from_data=True)
            chart.set_categories(categories_reference)

            # Добавление графика на лист
            graph_sheet.add_chart(chart, f"A{start_row}")

            start_row += len(data_sheet1) + 10
            chart_title = "График 2"
            chart_type = LineChart
            data_reference = Reference(workbook["лр1"], min_col=4, min_row=1, max_col=4, max_row=len(data_sheet2) + 1)
            categories_reference = Reference(workbook["лр1"], min_col=1, min_row=2, max_row=len(data_sheet2) + 1)

            chart = chart_type()
            chart.title = chart_title
            chart.x_axis.title = 'ФИО'
            chart.y_axis.title = 'Значение'
            chart.add_data(data_reference, titles_from_data=True)
            chart.set_categories(categories_reference)
            graph_sheet.add_chart(chart, f"A{start_row}")

            start_row += len(data_sheet2) + 10
            self.create_chart("График 3", AreaChart,
                         Reference(workbook["Баллы"], min_col=2, min_row=1, max_col=2, max_row=len(data_sheet3) + 1),
                         Reference(workbook["Баллы"], min_col=1, min_row=2, max_row=len(data_sheet3) + 1), graph_sheet,
                         start_row)

            # Сохранение обновленной книги
            workbook.save(self.file_path)
            print(f"Графики успешно сохранены в файле: {self.file_path}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
