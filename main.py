import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
from PyQt5.QtWidgets import *


class MyWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.file_path = ''
        self.search_name = ''
        self.init_ui()

    def init_ui(self) -> None:
        create = QPushButton("Create schedule")
        search = QPushButton("Open workbook")
        cental_widget = QWidget()
        self.create_line_edit = QLineEdit()
        self.create_line_edit.setFixedSize(300, 20)
        self.search_label = QLabel()
        self.search_label.setFixedSize(300, 20)
        self.search_label.setStyleSheet("background: white")
        self.setCentralWidget(cental_widget)
        main_layout = QVBoxLayout()
        hbox = QHBoxLayout()
        hbox.addWidget(create)
        hbox.addWidget(self.create_line_edit)
        main_layout.addLayout(hbox)
        hbox = QHBoxLayout()
        hbox.addWidget(search)
        hbox.addWidget(self.search_label)
        create.clicked.connect(self.highlight_cells_xls)
        search.clicked.connect(self.open_xls)
        main_layout.addLayout(hbox)
        self.short_version = QCheckBox("Short version")
        main_layout.addWidget(self.short_version)
        cental_widget.setLayout(main_layout)
        self.setFixedSize(410, 93)

    def open_xls(self) -> None:
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(self,
                                                   'Open Excel File ',
                                                   '',
                                                   'Excel File (*.xlsx)',
                                                   options=options)
        if file_name:
            self.file_path = file_name
            self.search_label.setText(file_name)

    def highlight_cells_xls(self) -> None:
        self.search_name = self.create_line_edit.text()
        if not self.search_name or not self.file_path:
            return
        wb = load_workbook(self.file_path)
        week = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]
        first_week = {i: [] for i in week}
        second_week = {i: [] for i in week}
        for sheet in wb.sheetnames:
            wb_sheet = wb[sheet]
            days = {}
            for i in range(5, len(wb_sheet['B'])):
                if wb_sheet[f'B{i}'].value is not None:
                    days[i] = wb_sheet[f'B{i}'].value
            days_keys = [key for key in days.keys()]
            last_key = days_keys[1]
            key = days_keys[0]
            group = wb_sheet["A5"].value
            for i in range(5, len(wb_sheet['G'])):
                if last_key <= i and last_key != days_keys[-1]:
                    if days[key] == "ПН":
                        group = wb_sheet[f"A{key}"].value
                    key = last_key
                    last_key = days_keys[days_keys.index(last_key) + 1]
                if self.search_name in str(wb_sheet[f"G{i}"].value):
                    first_week[days[key]].append([group])
                    for j in "CDEFGH":
                        first_week[days[key]][-1].append(wb_sheet[f"{j}{i}"].value)
                if self.search_name in str(wb_sheet[f"N{i}"].value):
                    second_week[days[key]].append([group])
                    for j in "JKLMNO":
                        second_week[days[key]][-1].append(wb_sheet[f"{j}{i}"].value)

        wb = Workbook()
        ws = wb.active
        ws.title = self.create_line_edit.text()
        thin = Side(border_style="thin", color="000000")
        letters = "ABCDEFGHIJKLMNOP"
        words = """День недели 
        Группа 
        № Пары 
        Предмет 
        Тип занятия 
        Подгруппа 
        Преподаватель 
        Аудитория 
        День недели 
        Группа 
        № Пары 
        Предмет 
        Тип занятия 
        Подгруппа 
        Преподаватель 
        Аудитория"""
        words = [i for i in words.split('\n')]
        for i in range(len(letters)):
            ws[f"{letters[i]}4"].value = words[i]
        for i in range(5, 7 * 5 + 6):
            if i % 6 == 0:
                for j in ws[f"{i - 1}"]:
                    j.border = Border(top=thin)
                ws[f"A{i - 1}"].value = week[i // 6 - 1]
                ws[f"I{i - 1}"].value = week[i // 6 - 1]
                for j in 'ABCHIJKP':
                    ws[f"{j}{i - 1}"].border = Border(right=thin, left=thin, top=thin)
            else:
                for j in 'ABCHIJKP':
                    ws[f"{j}{i - 1}"].border = Border(right=thin, left=thin)
            for j in first_week[week[i // 6 - 1]]:
                if i % 6 - 1 in j:
                    for z, k in enumerate("BCDFEGH"):
                        ws[f"{k}{i - 3}"].value = j[z]
            for j in second_week[week[i // 6 - 1]]:
                if i % 6 - 1 in j:
                    for z, k in enumerate("JKLNMOP"):
                        ws[f"{k}{i - 3}"].value = j[z]
        for i in ws[f"3"]:
            i.border = Border(bottom=thin)
        for i in "DGLO":
            ws.column_dimensions[f"{i}"].width = 35

        if self.short_version.checkState():
            for k in range(5):
                for i in range(5, 7 * 5 + 6):
                    is_clear = True
                    for j in ws[f"{i}"]:
                        if j.value is not None:
                            is_clear = False
                            break
                    if is_clear:
                        ws.delete_rows(i)

        for i in ws[f"{len(ws['A']) + 1}"]:
            i.border = Border(top=thin)
        wb.save(self.search_name + ".xlsx")
        sys.exit()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
